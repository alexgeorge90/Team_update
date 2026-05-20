"""
Archive closed tasks: reads team_data.xlsx, moves any task whose latest
Weekly Update status is "Closed" from the Tasks tab to the Archived Tasks tab,
preserving the original Task ID.  Run this whenever you mark a task Closed.

Usage:  python archive_tasks.py
"""

import sys
from copy import copy
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_FILE = "sample_data/team_data.xlsx"

ARCHIVE_HEADER_FILL = PatternFill(start_color="805AD5", end_color="805AD5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

TASK_HEADERS = [
    "Task Name", "Description", "Team Members",
    "Start Date", "Target End Date", "Status", "Task ID",
]


def resolve_status(wb, task_name):
    """Find the latest status for a task from the Weekly Updates sheet."""
    ws = wb["Weekly Updates"]
    latest_date = None
    latest_status = None
    for row in ws.iter_rows(min_row=2, max_col=8, values_only=False):
        cell_task = row[1].value  # col B = Task
        cell_status = row[2].value  # col C = Status
        cell_date = row[7].value  # col H = Update Date
        if cell_task == task_name and cell_status:
            if isinstance(cell_date, datetime):
                dt = cell_date
            elif cell_date:
                try:
                    dt = datetime.strptime(str(cell_date), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    dt = datetime.min
            else:
                dt = datetime.min
            if latest_date is None or dt >= latest_date:
                latest_date = dt
                latest_status = cell_status
    return latest_status


def ensure_archive_sheet(wb):
    if "Archived Tasks" in wb.sheetnames:
        return wb["Archived Tasks"]
    ws = wb.create_sheet("Archived Tasks")
    ws.sheet_properties.tabColor = "805AD5"
    for col, header in enumerate(TASK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = ARCHIVE_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["G"].width = 12
    ws.freeze_panes = "A2"
    return ws


def next_archive_row(ws):
    row = 2
    while ws.cell(row=row, column=1).value:
        row += 1
    return row


def archive_closed(path=DATA_FILE):
    wb = load_workbook(path)
    ws_tasks = wb["Tasks"]
    ws_archive = ensure_archive_sheet(wb)

    archived = []
    rows_to_delete = []

    for row in range(2, ws_tasks.max_row + 1):
        task_name = ws_tasks.cell(row=row, column=1).value
        if not task_name:
            continue

        status = resolve_status(wb, task_name)
        if status != "Closed":
            continue

        task_id = ws_tasks.cell(row=row, column=7).value
        # If Task ID is a formula, compute it from the row position
        if isinstance(task_id, str) and task_id.startswith("="):
            task_id = f"DRM{row - 1:03d}"

        dest_row = next_archive_row(ws_archive)
        ws_archive.cell(row=dest_row, column=1, value=task_name)
        ws_archive.cell(row=dest_row, column=2, value=ws_tasks.cell(row=row, column=2).value)
        ws_archive.cell(row=dest_row, column=3, value=ws_tasks.cell(row=row, column=3).value)

        start_val = ws_tasks.cell(row=row, column=4).value
        ws_archive.cell(row=dest_row, column=4, value=start_val)
        ws_archive.cell(row=dest_row, column=4).number_format = "DD-MMM-YYYY"

        end_val = ws_tasks.cell(row=row, column=5).value
        ws_archive.cell(row=dest_row, column=5, value=end_val)
        ws_archive.cell(row=dest_row, column=5).number_format = "DD-MMM-YYYY"

        ws_archive.cell(row=dest_row, column=6, value="Closed")
        ws_archive.cell(row=dest_row, column=7, value=task_id)

        for col in range(1, 8):
            ws_archive.cell(row=dest_row, column=col).border = THIN_BORDER
            ws_archive.cell(row=dest_row, column=col).alignment = Alignment(
                vertical="center", wrap_text=True,
            )

        rows_to_delete.append(row)
        archived.append(f"  {task_id} : {task_name}")

    if not archived:
        print("No closed tasks to archive.")
        wb.close()
        return

    # Delete rows from Tasks tab bottom-up so row indices stay valid
    for row in reversed(rows_to_delete):
        ws_tasks.delete_rows(row, 1)

    wb.save(path)
    wb.close()
    print(f"Archived {len(archived)} task(s):")
    for line in archived:
        print(line)
    print(f"\nSaved to {path}")


if __name__ == "__main__":
    archive_closed()
