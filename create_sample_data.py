"""
Generate a validated Excel workbook with 5 tabs:
  - Tasks          : active tasks only (Task ID = DRM001 as last col)
  - Archived Tasks : closed tasks moved here (preserves original Task ID)
  - Weekly Updates  : task dropdown from Tasks tab only, renamed cols
  - Allocations    : task dropdown, team member dropdown
  - Lookups        : reference lists
"""

import os
from datetime import date
from itertools import combinations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

os.makedirs("sample_data", exist_ok=True)

MEMBERS = ["Alex", "Srikanth", "Enith"]
STATUSES = ["Not Started", "In Progress", "On Track", "At Risk", "Blocked", "Closed"]
UPDATE_STATUSES = ["On Track", "At Risk", "Blocked", "Closed"]
MAX_TASK_ROWS = 50
MAX_UPDATE_ROWS = 500
MAX_ALLOC_ROWS = 200

MEMBER_COMBOS = []
for r in range(1, len(MEMBERS) + 1):
    for combo in combinations(MEMBERS, r):
        MEMBER_COMBOS.append("; ".join(combo))

# ── Styling ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ARCHIVE_HEADER_FILL = PatternFill(start_color="805AD5", end_color="805AD5", fill_type="solid")
LOOKUP_HEADER_FILL = PatternFill(start_color="718096", end_color="718096", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def style_header(ws, col_count, fill=None):
    fill = fill or HEADER_FILL
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_area(ws, max_row, col_count, locked_cols=None):
    locked_cols = locked_cols or []
    for row in range(2, max_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col in locked_cols:
                cell.fill = LOCKED_FILL


def auto_width(ws, col_count, min_width=12, max_width=30):
    for col in range(1, col_count + 1):
        header_len = len(str(ws.cell(row=1, column=col).value or ""))
        width = max(min_width, min(header_len + 4, max_width))
        ws.column_dimensions[get_column_letter(col)].width = width


# ── Sample data ──────────────────────────────────────────────────────────────

ALL_TASKS = [
    ("API Gateway Build", "Build and deploy API gateway with rate limiting and auth",
     "Alex; Srikanth", date(2026, 1, 15), date(2026, 6, 30)),
    ("Dashboard UI", "Frontend dashboard with charts and filters",
     "Alex; Enith", date(2026, 3, 1), date(2026, 7, 31)),
    ("Auth Module", "OIDC-based authentication module",
     "Srikanth", date(2026, 3, 15), date(2026, 5, 31)),
    ("Perf Testing", "Load and stress testing across all services",
     "Alex; Srikanth; Enith", date(2026, 5, 1), date(2026, 8, 31)),
    ("Data Migration", "Migrate legacy DB to new schema",
     "Enith", date(2026, 2, 1), date(2026, 4, 30)),
]

week_dates = [
    date(2026, 3, 23), date(2026, 3, 30),
    date(2026, 4, 6), date(2026, 4, 13),
    date(2026, 4, 20), date(2026, 4, 27),
    date(2026, 5, 4), date(2026, 5, 11),
]

member_task_data = {
    "Alex": [
        ("API Gateway Build", [18, 20, 22, 20, 15, 18, 20, 15]),
        ("Dashboard UI", [0, 0, 0, 5, 10, 12, 12, 15]),
    ],
    "Srikanth": [
        ("API Gateway Build", [20, 22, 18, 15, 12, 10, 8, 5]),
        ("Auth Module", [15, 18, 20, 22, 25, 25, 28, 30]),
    ],
    "Enith": [
        ("Data Migration", [35, 32, 30, 25, 10, 5, 0, 0]),
        ("Dashboard UI", [0, 5, 8, 12, 25, 30, 32, 35]),
    ],
}

statuses_by_task = {
    "API Gateway Build": ["On Track"] * 8,
    "Dashboard UI": ["On Track"] * 4 + ["At Risk"] * 4,
    "Auth Module": ["On Track"] * 3 + ["At Risk"] * 5,
    "Data Migration": ["On Track"] * 4 + ["On Track", "Closed", "Closed", "Closed"],
}

progress_notes = {
    "API Gateway Build": [
        "Route config done", "Auth middleware started", "Rate limiter built",
        "Integration tests", "Load testing prep", "API docs written",
        "Staging deploy", "Load testing started",
    ],
    "Dashboard UI": [
        "", "", "", "Initial wireframes",
        "Chart components built", "Filter panel done", "Data binding complete",
        "Started user testing",
    ],
    "Auth Module": [
        "OIDC research", "Provider setup", "Token flow built",
        "OIDC integration issues", "Vendor delay on certs",
        "Workaround in progress", "Partial integration done", "Blocked on vendor",
    ],
    "Data Migration": [
        "Schema mapping done", "ETL scripts written", "Test migration run",
        "Data validation", "Final migration", "Cutover complete", "", "",
    ],
}

dependencies_data = {
    "Auth Module": ["", "", "", "Vendor delay", "Vendor delay",
                    "Vendor delay", "Vendor delay", "Vendor delay"],
}

next_actions = {
    "API Gateway Build": [
        "Start auth middleware", "Build rate limiter", "Integration tests",
        "Prep load testing", "Write API docs", "Deploy to staging",
        "Start load testing", "Analyze results",
    ],
    "Dashboard UI": [
        "", "", "", "Build chart components",
        "Build filter panel", "Data binding", "User testing",
        "Fix feedback items",
    ],
    "Auth Module": [
        "Provider setup", "Build token flow", "OIDC integration",
        "Follow up with vendor", "Explore workaround",
        "Continue workaround", "Integration testing", "Resolve vendor block",
    ],
    "Data Migration": [
        "Write ETL scripts", "Test migration", "Data validation",
        "Final migration", "Cutover", "Post-cutover monitoring", "", "",
    ],
}

# Determine each task's latest status from the weekly updates
def get_latest_status(task_name):
    for member, tasks_list in member_task_data.items():
        for task, hours_list in tasks_list:
            if task == task_name:
                statuses = statuses_by_task.get(task, [])
                for s in reversed(statuses):
                    if s:
                        return s
    return "Not Started"

# Split tasks into active and archived based on latest status
active_tasks = []
archived_tasks = []
for idx, task_tuple in enumerate(ALL_TASKS, start=1):
    name = task_tuple[0]
    latest = get_latest_status(name)
    task_id = f"DRM{idx:03d}"
    if latest == "Closed":
        archived_tasks.append((task_id, *task_tuple, latest))
    else:
        active_tasks.append(task_tuple)


# ── Build workbook ───────────────────────────────────────────────────────────

wb = Workbook()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB: Lookups
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ws_lookup = wb.active
ws_lookup.title = "Lookups"
ws_lookup.sheet_properties.tabColor = "718096"

ws_lookup["A1"] = "Team Members"
for i, combo in enumerate(MEMBER_COMBOS, start=2):
    ws_lookup.cell(row=i, column=1, value=combo)

ws_lookup["B1"] = "Members"
for i, m in enumerate(MEMBERS, start=2):
    ws_lookup.cell(row=i, column=2, value=m)

ws_lookup["C1"] = "Task Statuses"
for i, s in enumerate(STATUSES, start=2):
    ws_lookup.cell(row=i, column=3, value=s)

ws_lookup["D1"] = "Update Statuses"
for i, s in enumerate(UPDATE_STATUSES, start=2):
    ws_lookup.cell(row=i, column=4, value=s)

style_header(ws_lookup, 4, fill=LOOKUP_HEADER_FILL)
auto_width(ws_lookup, 4, min_width=18)

# TaskList: dynamic range on Tasks!$A, auto-sizes to non-blank rows (no closed tasks here)
wb.defined_names.add(DefinedName(
    "TaskList",
    attr_text=f"OFFSET(Tasks!$A$2,0,0,COUNTA(Tasks!$A$2:$A${MAX_TASK_ROWS + 1}),1)",
))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB: Tasks  (active only -- closed tasks go to Archived Tasks)
# Columns: Task Name, Description, Team Members, Start Date, Target End Date,
#          Status (formula), Task ID (DRM format, last col)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ws_tasks = wb.create_sheet("Tasks", 0)
ws_tasks.sheet_properties.tabColor = "2B6CB0"

task_headers = [
    "Task Name", "Description", "Team Members",
    "Start Date", "Target End Date", "Status", "Task ID",
]
for col, header in enumerate(task_headers, start=1):
    ws_tasks.cell(row=1, column=col, value=header)

# Task ID formula (col G = 7): sequential DRM001, DRM002... based on row.
# Active tasks get clean sequential IDs; archived tasks keep their original IDs.
for row in range(2, MAX_TASK_ROWS + 2):
    ws_tasks.cell(
        row=row, column=7,
        value=f'=IF(A{row}="","","DRM"&TEXT(ROW()-1,"000"))',
    )

# Status formula (col F = 6): auto-picks last status from Weekly Updates
# Weekly Updates: col B = Task, col C = Status
end = MAX_UPDATE_ROWS + 1
for row in range(2, MAX_TASK_ROWS + 2):
    formula = (
        f'=IF(A{row}="","",'
        f"IFERROR(LOOKUP(2,1/('Weekly Updates'!$B$2:$B${end}=A{row}),"
        f"'Weekly Updates'!$C$2:$C${end})"
        f',"Not Started"))'
    )
    ws_tasks.cell(row=row, column=6, value=formula)

# Data validation: Team Members dropdown (col C = 3)
combo_str = ",".join(MEMBER_COMBOS)
dv_team = DataValidation(
    type="list", formula1=f'"{combo_str}"',
    showErrorMessage=True,
    errorTitle="Invalid Team",
    error="Select a valid team member combination from the dropdown.",
)
dv_team.add(f"C2:C{MAX_TASK_ROWS + 1}")
ws_tasks.add_data_validation(dv_team)

# Date format for Start Date (col D) & Target End Date (col E)
for row in range(2, MAX_TASK_ROWS + 2):
    ws_tasks.cell(row=row, column=4).number_format = "DD-MMM-YYYY"
    ws_tasks.cell(row=row, column=5).number_format = "DD-MMM-YYYY"

style_header(ws_tasks, len(task_headers))
style_data_area(ws_tasks, MAX_TASK_ROWS + 1, len(task_headers), locked_cols=[6, 7])
auto_width(ws_tasks, len(task_headers))
ws_tasks.column_dimensions["A"].width = 22
ws_tasks.column_dimensions["B"].width = 35
ws_tasks.column_dimensions["C"].width = 25
ws_tasks.column_dimensions["G"].width = 12
ws_tasks.freeze_panes = "A2"

# Write active sample tasks
for i, (name, desc, team, start, end_dt) in enumerate(active_tasks, start=2):
    ws_tasks.cell(row=i, column=1, value=name)
    ws_tasks.cell(row=i, column=2, value=desc)
    ws_tasks.cell(row=i, column=3, value=team)
    ws_tasks.cell(row=i, column=4, value=start)
    ws_tasks.cell(row=i, column=5, value=end_dt)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB: Archived Tasks  (closed tasks land here with their original Task IDs)
# Same columns as Tasks, but Task ID is a stored value, Status is "Closed"
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ws_archive = wb.create_sheet("Archived Tasks", 1)
ws_archive.sheet_properties.tabColor = "805AD5"

archive_headers = [
    "Task Name", "Description", "Team Members",
    "Start Date", "Target End Date", "Status", "Task ID",
]
for col, header in enumerate(archive_headers, start=1):
    ws_archive.cell(row=1, column=col, value=header)

style_header(ws_archive, len(archive_headers), fill=ARCHIVE_HEADER_FILL)

for row in range(2, 202):
    ws_archive.cell(row=row, column=4).number_format = "DD-MMM-YYYY"
    ws_archive.cell(row=row, column=5).number_format = "DD-MMM-YYYY"

style_data_area(ws_archive, 20, len(archive_headers))
auto_width(ws_archive, len(archive_headers))
ws_archive.column_dimensions["A"].width = 22
ws_archive.column_dimensions["B"].width = 35
ws_archive.column_dimensions["C"].width = 25
ws_archive.column_dimensions["G"].width = 12
ws_archive.freeze_panes = "A2"

# Write archived tasks (static values -- no formulas)
for i, (task_id, name, desc, team, start, end_dt, status) in enumerate(archived_tasks, start=2):
    ws_archive.cell(row=i, column=1, value=name)
    ws_archive.cell(row=i, column=2, value=desc)
    ws_archive.cell(row=i, column=3, value=team)
    ws_archive.cell(row=i, column=4, value=start)
    ws_archive.cell(row=i, column=5, value=end_dt)
    ws_archive.cell(row=i, column=6, value=status)
    ws_archive.cell(row=i, column=7, value=task_id)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB: Weekly Updates
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ws_updates = wb.create_sheet("Weekly Updates", 2)
ws_updates.sheet_properties.tabColor = "38A169"

update_headers = [
    "Team Member", "Task", "Status", "Progress_Notes",
    "Dependency", "Planned_Next_Action", "Hours Spent", "Update Date",
]
for col, header in enumerate(update_headers, start=1):
    ws_updates.cell(row=1, column=col, value=header)

member_str = ",".join(MEMBERS)
dv_member = DataValidation(
    type="list", formula1=f'"{member_str}"',
    showErrorMessage=True,
    errorTitle="Invalid Member",
    error="Select a team member from the dropdown.",
)
dv_member.add(f"A2:A{MAX_UPDATE_ROWS + 1}")
ws_updates.add_data_validation(dv_member)

# Task dropdown: uses dynamic named range from Tasks tab (only active tasks)
dv_task = DataValidation(
    type="list", formula1="=TaskList",
    showErrorMessage=True,
    errorTitle="Invalid Task",
    error="Select a task from the dropdown.",
)
dv_task.add(f"B2:B{MAX_UPDATE_ROWS + 1}")
ws_updates.add_data_validation(dv_task)

update_status_str = ",".join(UPDATE_STATUSES)
dv_update_status = DataValidation(
    type="list", formula1=f'"{update_status_str}"',
    showErrorMessage=True,
    errorTitle="Invalid Status",
    error="Select a status from the dropdown.",
)
dv_update_status.add(f"C2:C{MAX_UPDATE_ROWS + 1}")
ws_updates.add_data_validation(dv_update_status)

dv_hours = DataValidation(
    type="whole", operator="between",
    formula1="0", formula2="80",
    showErrorMessage=True,
    errorTitle="Invalid Hours",
    error="Hours must be between 0 and 80.",
)
dv_hours.add(f"G2:G{MAX_UPDATE_ROWS + 1}")
ws_updates.add_data_validation(dv_hours)

for row in range(2, MAX_UPDATE_ROWS + 2):
    ws_updates.cell(row=row, column=8).number_format = "DD-MMM-YYYY"

style_header(ws_updates, len(update_headers))
style_data_area(ws_updates, 50, len(update_headers), locked_cols=[8])
auto_width(ws_updates, len(update_headers))
ws_updates.column_dimensions["D"].width = 30
ws_updates.column_dimensions["E"].width = 20
ws_updates.column_dimensions["F"].width = 25
ws_updates.freeze_panes = "A2"

# Write sample update data (all tasks, including ones now archived -- historical data stays)
row_num = 2
for member, tasks_list in member_task_data.items():
    for task, hours_list in tasks_list:
        for i, (wk, hrs) in enumerate(zip(week_dates, hours_list)):
            if hrs > 0:
                ws_updates.cell(row=row_num, column=1, value=member)
                ws_updates.cell(row=row_num, column=2, value=task)
                ws_updates.cell(row=row_num, column=3, value=statuses_by_task[task][i])
                ws_updates.cell(row=row_num, column=4, value=progress_notes[task][i])
                dep = dependencies_data.get(task, [""] * 8)
                ws_updates.cell(row=row_num, column=5, value=dep[i] if i < len(dep) else "")
                na = next_actions.get(task, [""] * 8)
                ws_updates.cell(row=row_num, column=6, value=na[i] if i < len(na) else "")
                ws_updates.cell(row=row_num, column=7, value=hrs)
                ws_updates.cell(row=row_num, column=8, value=wk)
                row_num += 1

total_updates = row_num - 2

for row in range(row_num, MAX_UPDATE_ROWS + 2):
    ws_updates.cell(
        row=row, column=8,
        value=f'=IF(A{row}="","",TODAY())',
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TAB: Allocations
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ws_alloc = wb.create_sheet("Allocations", 3)
ws_alloc.sheet_properties.tabColor = "DD6B20"

alloc_headers = ["Week Starting", "Team Member", "Task", "Planned Hours"]
for col, header in enumerate(alloc_headers, start=1):
    ws_alloc.cell(row=1, column=col, value=header)

dv_alloc_member = DataValidation(
    type="list", formula1=f'"{member_str}"',
    showErrorMessage=True,
    errorTitle="Invalid Member",
    error="Select a team member from the dropdown.",
)
dv_alloc_member.add(f"B2:B{MAX_ALLOC_ROWS + 1}")
ws_alloc.add_data_validation(dv_alloc_member)

dv_alloc_task = DataValidation(
    type="list", formula1="=TaskList",
    showErrorMessage=True,
    errorTitle="Invalid Task",
    error="Select a task from the dropdown.",
)
dv_alloc_task.add(f"C2:C{MAX_ALLOC_ROWS + 1}")
ws_alloc.add_data_validation(dv_alloc_task)

dv_plan_hours = DataValidation(
    type="whole", operator="between",
    formula1="0", formula2="60",
    showErrorMessage=True,
    errorTitle="Invalid Hours",
    error="Planned hours must be between 0 and 60.",
)
dv_plan_hours.add(f"D2:D{MAX_ALLOC_ROWS + 1}")
ws_alloc.add_data_validation(dv_plan_hours)

for row in range(2, MAX_ALLOC_ROWS + 2):
    ws_alloc.cell(row=row, column=1).number_format = "DD-MMM-YYYY"

style_header(ws_alloc, len(alloc_headers))
style_data_area(ws_alloc, 50, len(alloc_headers))
auto_width(ws_alloc, len(alloc_headers), min_width=16)
ws_alloc.freeze_panes = "A2"

future_weeks = [
    date(2026, 5, 18), date(2026, 5, 25),
    date(2026, 6, 1), date(2026, 6, 8),
]

alloc_plan = {
    "Alex": [
        ("API Gateway Build", [15, 10, 5, 0]),
        ("Dashboard UI", [20, 20, 25, 25]),
        ("Perf Testing", [0, 5, 10, 15]),
    ],
    "Srikanth": [
        ("API Gateway Build", [5, 5, 0, 0]),
        ("Auth Module", [30, 25, 15, 10]),
        ("Perf Testing", [0, 10, 20, 25]),
    ],
    "Enith": [
        ("Dashboard UI", [30, 25, 15, 10]),
        ("Perf Testing", [5, 10, 20, 25]),
    ],
}

row_num = 2
for member, plan_list in alloc_plan.items():
    for task, hours_list in plan_list:
        for wk, hrs in zip(future_weeks, hours_list):
            if hrs > 0:
                ws_alloc.cell(row=row_num, column=1, value=wk)
                ws_alloc.cell(row=row_num, column=2, value=member)
                ws_alloc.cell(row=row_num, column=3, value=task)
                ws_alloc.cell(row=row_num, column=4, value=hrs)
                row_num += 1

total_allocs = row_num - 2

# Move Lookups to the end
wb.move_sheet(ws_lookup, offset=4)

# Save
output_path = "sample_data/team_data.xlsx"
wb.save(output_path)
print(f"Created {output_path}")
print(f"  Tasks tab        : {len(active_tasks)} active tasks")
print(f"  Archived Tasks   : {len(archived_tasks)} closed tasks (preserved IDs)")
print(f"  Weekly Updates   : {total_updates} entries")
print(f"  Allocations      : {total_allocs} entries")
print(f"  Lookups tab      : reference lists")
